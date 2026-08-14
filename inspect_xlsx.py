import os
from openpyxl import load_workbook

path = r'C:\Users\TEDDYPAKER\Desktop\FindHive Repo\findhive\pokemon cardz.xlsx'
print('exists:', os.path.exists(path))
wb = load_workbook(path, read_only=True, data_only=True)
print('sheets:', wb.sheetnames)
for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    print(f'\nSHEET: {ws.title} rows={len(rows)} cols={ws.max_column}')
    for idx, row in enumerate(rows[:12], start=1):
        print(f'ROW {idx}: {row}')
